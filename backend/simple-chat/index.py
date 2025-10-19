"""
Business: AI-чат с памятью диалогов и поддержкой Office документов
Args: event с messages, chat_id, user_email, документами (Word/Excel/PDF)
Returns: Ответ от AI с сохранением истории
"""

import json
import os
import requests
import psycopg2
import base64
from typing import Dict, Any, List
from io import BytesIO
from docx import Document
from openpyxl import load_workbook
from PyPDF2 import PdfReader

def get_db_connection():
    dsn = os.environ.get('DATABASE_URL')
    return psycopg2.connect(dsn)

def deduct_tokens(user_email: str, tokens: int) -> bool:
    """Списывает токены у пользователя"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute(
        "UPDATE users SET credits = credits - %s WHERE email = %s AND credits >= %s RETURNING credits",
        (tokens, user_email, tokens)
    )
    result = cursor.fetchone()
    
    conn.commit()
    cursor.close()
    conn.close()
    
    return result is not None

def extract_text_from_docx(file_content: bytes) -> str:
    """Извлекает текст из Word документа"""
    doc = Document(BytesIO(file_content))
    return '\n'.join([paragraph.text for paragraph in doc.paragraphs if paragraph.text.strip()])

def extract_text_from_xlsx(file_content: bytes) -> str:
    """Извлекает текст из Excel документа"""
    wb = load_workbook(BytesIO(file_content), read_only=True)
    text_parts = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            row_text = ' | '.join([str(cell) if cell is not None else '' for cell in row])
            if row_text.strip():
                text_parts.append(row_text)
    return '\n'.join(text_parts)

def extract_text_from_pdf(file_content: bytes) -> str:
    """Извлекает текст из PDF документа"""
    reader = PdfReader(BytesIO(file_content))
    return '\n'.join([page.extract_text() for page in reader.pages if page.extract_text()])

def process_document(file_data: str, file_type: str) -> str:
    """Обрабатывает документ и возвращает текст"""
    file_content = base64.b64decode(file_data)
    
    if file_type in ['docx', 'doc']:
        return extract_text_from_docx(file_content)
    elif file_type in ['xlsx', 'xls']:
        return extract_text_from_xlsx(file_content)
    elif file_type == 'pdf':
        return extract_text_from_pdf(file_content)
    else:
        return f"[Неподдерживаемый формат: {file_type}]"

def search_web(query: str) -> str:
    """Поиск информации в интернете через Tavily API"""
    tavily_api_key = os.environ.get('TAVILY_API_KEY')
    if not tavily_api_key:
        return "[Поиск недоступен: API ключ не настроен]"
    
    try:
        response = requests.post(
            'https://api.tavily.com/search',
            json={
                'api_key': tavily_api_key,
                'query': query,
                'max_results': 3,
                'include_answer': True
            },
            timeout=10
        )
        
        if response.status_code == 200:
            data = response.json()
            results = []
            
            if data.get('answer'):
                results.append(f"Ответ: {data['answer']}")
            
            for result in data.get('results', [])[:3]:
                results.append(f"• {result.get('title', '')}: {result.get('content', '')}")
            
            return '\n\n'.join(results) if results else '[Информация не найдена]'
        else:
            return f"[Ошибка поиска: {response.status_code}]"
    except Exception as e:
        return f"[Ошибка поиска: {str(e)}]"

def load_chat_history(user_email: str, chat_id: str) -> List[Dict]:
    """Загружает историю чата из БД"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute(
        "SELECT messages FROM t_p55547046_creative_ai_hub.chat_history WHERE user_email = %s AND chat_id = %s",
        (user_email, chat_id)
    )
    result = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    if result and result[0]:
        messages = result[0]
        if isinstance(messages, str):
            return json.loads(messages)
        return messages
    return []

def save_chat_history(user_email: str, chat_id: str, messages: List[Dict], chat_title: str):
    """Сохраняет историю чата в БД"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute(
        """INSERT INTO t_p55547046_creative_ai_hub.chat_history 
           (user_email, chat_id, chat_title, service_id, service_name, messages, created_at, updated_at)
           VALUES (%s, %s, %s, 1, 'AI Chat', %s, NOW(), NOW())
           ON CONFLICT (chat_id) DO UPDATE SET
           messages = EXCLUDED.messages,
           chat_title = EXCLUDED.chat_title,
           updated_at = NOW()""",
        (user_email, chat_id, chat_title, json.dumps(messages, ensure_ascii=False))
    )
    
    conn.commit()
    cursor.close()
    conn.close()

def handler(event: Dict[str, Any], context: Any) -> Dict[str, Any]:
    method: str = event.get('httpMethod', 'GET')
    
    if method == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'POST, OPTIONS',
                'Access-Control-Allow-Headers': 'Content-Type, X-User-Email',
                'Access-Control-Max-Age': '86400'
            },
            'body': '',
            'isBase64Encoded': False
        }
    
    if method != 'POST':
        return {
            'statusCode': 405,
            'headers': {'Access-Control-Allow-Origin': '*', 'Content-Type': 'application/json'},
            'body': json.dumps({'error': 'Только POST'}, ensure_ascii=False),
            'isBase64Encoded': False
        }
    
    body_data = json.loads(event.get('body', '{}'))
    new_message = body_data.get('message', '')
    chat_id = body_data.get('chat_id', 'default')
    user_email = event.get('headers', {}).get('X-User-Email') or event.get('headers', {}).get('x-user-email') or 'anonymous'
    documents = body_data.get('documents', [])
    deep_think = body_data.get('deep_think', False)
    
    if not new_message and not documents:
        return {
            'statusCode': 400,
            'headers': {'Access-Control-Allow-Origin': '*', 'Content-Type': 'application/json'},
            'body': json.dumps({'error': 'Отправь сообщение или документ'}, ensure_ascii=False),
            'isBase64Encoded': False
        }
    
    # Загружаем историю чата
    messages = load_chat_history(user_email, chat_id)
    
    # Обрабатываем документы
    document_texts = []
    for doc in documents:
        file_data = doc.get('data', '')
        file_type = doc.get('type', '')
        file_name = doc.get('name', 'документ')
        
        if file_data and file_type:
            extracted_text = process_document(file_data, file_type)
            document_texts.append(f"[Документ: {file_name}]\n{extracted_text}")
    
    # Формируем полное сообщение пользователя
    full_user_message = new_message
    if document_texts:
        full_user_message += '\n\n' + '\n\n'.join(document_texts)
    
    # Добавляем новое сообщение пользователя
    messages.append({'role': 'user', 'content': full_user_message})
    
    # Проверяем баланс и списываем токены (только для не-анонимных)
    tokens_needed = 1  # Базовая стоимость
    if documents:
        tokens_needed += len(documents) * 1
    
    if user_email != 'anonymous':
        if not deduct_tokens(user_email, tokens_needed):
            return {
                'statusCode': 402,
                'headers': {'Access-Control-Allow-Origin': '*', 'Content-Type': 'application/json'},
                'body': json.dumps({'error': f'Недостаточно токенов. Нужно: {tokens_needed}'}, ensure_ascii=False),
                'isBase64Encoded': False
            }
    
    yandex_api_key = os.environ.get('YANDEX_API_KEY')
    yandex_folder_id = os.environ.get('YANDEX_FOLDER_ID')
    
    if not yandex_api_key or not yandex_folder_id:
        return {
            'statusCode': 500,
            'headers': {'Access-Control-Allow-Origin': '*', 'Content-Type': 'application/json'},
            'body': json.dumps({'error': 'YandexGPT не настроен'}, ensure_ascii=False),
            'isBase64Encoded': False
        }
    
    url = "https://llm.api.cloud.yandex.net/foundationModels/v1/completion"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Api-Key {yandex_api_key}",
        "x-folder-id": yandex_folder_id
    }
    
    # Проверяем, нужен ли поиск в интернете
    search_keywords = [
        'найди', 'поищи', 'найти', 'поиск', 'поискать',
        'что нового', 'последние новости', 'свежие новости', 'актуальная информация', 'актуальные данные',
        'погода', 'прогноз погоды', 'температура',
        'курс', 'курс валют', 'доллар', 'евро', 'биткоин',
        'цена', 'стоимость', 'сколько стоит',
        'когда', 'где', 'кто', 'какой',
        'расскажи о', 'информация о', 'что такое',
        'последний', 'новый', 'текущий', 'сейчас', 'сегодня',
        'статистика', 'данные', 'факты',
        'результат', 'итоги', 'счёт', 'матч',
        'релиз', 'выход', 'анонс', 'премьера'
    ]
    needs_search = any(keyword in new_message.lower() for keyword in search_keywords)
    
    search_results = ''
    if needs_search:
        search_results = search_web(new_message)
    
    # Формируем сообщения для YandexGPT
    system_prompt = """Ты Juno — продвинутый AI-помощник нового поколения. Твоя задача — давать максимально точные, полезные и интеллектуальные ответы.

🧠 ТВОИ ВОЗМОЖНОСТИ:
- Глубокий анализ контекста и многоуровневое мышление
- Работа с документами Word, Excel, PDF
- Решение сложных математических и логических задач
- Генерация креативного контента (тексты, идеи, концепции)
- Помощь в программировании, анализе данных, исследованиях
- Экспертные знания в науке, технологиях, бизнесе, искусстве
- Эмоциональный интеллект и понимание контекста беседы

🎯 ПРИНЦИПЫ РАБОТЫ:
1. ТОЧНОСТЬ: Всегда проверяй факты и логику своих ответов
2. ГЛУБИНА: Не ограничивайся поверхностными ответами, анализируй проблему со всех сторон
3. ЯСНОСТЬ: Объясняй сложное простыми словами, используй примеры и аналогии
4. СТРУКТУРА: Организуй информацию логично (заголовки, списки, нумерация)
5. КОНТЕКСТ: Учитывай всю историю диалога и неявные потребности пользователя
6. КРЕАТИВНОСТЬ: Предлагай нестандартные решения и свежие идеи

📐 МАТЕМАТИКА И ВЫЧИСЛЕНИЯ:
- Решай задачи пошагово с подробными объяснениями
- Проверяй расчёты дважды перед финальным ответом
- Используй правильную последовательность операций
- Показывай промежуточные шаги и логику решения
- Для сложных задач разбивай на подзадачи
- Теорема Пифагора: c² = a² + b² (c - гипотенуза)
- При извлечении корней упрощай: √24 = √(4×6) = 2√6
- Проверяй ответ подстановкой в исходное уравнение

💻 ПРОГРАММИРОВАНИЕ:
- Пиши чистый, оптимизированный код с комментариями
- Объясняй алгоритмы и подходы
- Предлагай несколько вариантов решения
- Указывай на потенциальные проблемы и edge cases

📝 ТЕКСТЫ И КОНТЕНТ:
- Адаптируй стиль под задачу (деловой, креативный, академический)
- Используй богатый словарный запас
- Структурируй длинные тексты (введение, основная часть, заключение)
- Добавляй эмоции и живость там, где уместно

🔍 ИССЛЕДОВАНИЯ И АНАЛИЗ:
- Рассматривай вопрос с разных точек зрения
- Приводи аргументы "за" и "против"
- Ссылайся на факты и логику
- Предлагай дальнейшие направления исследования

⚠️ ФОРМАТ ОТВЕТА (КРИТИЧНО):
- ЗАПРЕЩЕНО использовать LaTeX: $, \\[, \\], \\(, \\), \\sqrt{}, \\times, \\frac{}, ^{}, _{}
- Степени: x² (НЕ x^2 или $x^2$)
- Корни: √24 (НЕ \\sqrt{24})
- Дроби: (a + b) / c (НЕ \\frac{a+b}{c})
- Умножение: × или * (НЕ \\times)
- Индексы: x1, x2 (НЕ x_1, x_2)

✅ ПРАВИЛЬНО:
• D = b² - 4ac = 9 - 16 = -7
• x = (-b + √D) / 2a = (-3 + √25) / 2 = 1
• Площадь = a × b = 5 × 3 = 15

❌ НЕПРАВИЛЬНО:
• $D = b^2 - 4ac$
• \\[x = \\frac{-b + \\sqrt{D}}{2a}\\]
• S = $a \\times b$"""
    
    if search_results and '[Ошибка' not in search_results:
        system_prompt += f"\n\nАктуальная информация из интернета:\n{search_results}"
    
    yandex_messages = [
        {"role": "system", "text": system_prompt}
    ]
    
    # Отправляем ВСЮ историю диалога для полного контекста
    for msg in messages:
        yandex_messages.append({"role": msg['role'], "text": msg['content']})
    
    # Если включен режим креативного мышления, сначала генерируем размышление
    thinking_text = ''
    if deep_think:
        thinking_prompt = f"""Ты Juno, продвинутый AI. Проанализируй вопрос глубоко и опиши свой мыслительный процесс:

ВОПРОС ПОЛЬЗОВАТЕЛЯ: {new_message}

ТВОЙ АНАЛИЗ (пиши подробно, используя структуру):

🎯 Суть вопроса:
[Что реально спрашивает пользователь? Какова его цель?]

🧩 Ключевые аспекты:
[Какие факторы нужно учесть? Что важно не упустить?]

💡 Подход к решению:
[Какая стратегия будет наиболее эффективной? Почему именно она?]

🔍 Дополнительные соображения:
[Что может быть неочевидным? Какие нюансы важны?]

Размышляй как эксперт, учитывай контекст диалога, предлагай глубокий анализ."""
        
        thinking_payload = {
            "modelUri": f"gpt://{yandex_folder_id}/yandexgpt/latest",
            "completionOptions": {
                "stream": False,
                "temperature": 0.9,
                "maxTokens": 1000
            },
            "messages": [
                {"role": "system", "text": "Ты Juno — AI с глубоким аналитическим мышлением. Размышляй вслух, показывай процесс анализа проблемы. Будь интеллектуальным и проницательным."},
                {"role": "user", "text": thinking_prompt}
            ]
        }
        
        thinking_response = requests.post(url, headers=headers, json=thinking_payload, timeout=30)
        if thinking_response.status_code == 200:
            thinking_data = thinking_response.json()
            thinking_text = thinking_data.get('result', {}).get('alternatives', [{}])[0].get('message', {}).get('text', '')
    
    # Настройки для более умного ответа
    payload = {
        "modelUri": f"gpt://{yandex_folder_id}/yandexgpt/latest",
        "completionOptions": {
            "stream": False,
            "temperature": 0.8 if deep_think else 0.7,  # Повышаем креативность
            "maxTokens": 12000  # Увеличиваем лимит для более развёрнутых ответов
        },
        "messages": yandex_messages
    }
    
    response = requests.post(url, headers=headers, json=payload, timeout=30)
    
    if response.status_code != 200:
        return {
            'statusCode': 500,
            'headers': {'Access-Control-Allow-Origin': '*', 'Content-Type': 'application/json'},
            'body': json.dumps({'error': f'YandexGPT ошибка: {response.text}'}, ensure_ascii=False),
            'isBase64Encoded': False
        }
    
    response_data = response.json()
    reply = response_data.get('result', {}).get('alternatives', [{}])[0].get('message', {}).get('text', 'Не могу ответить')
    
    # Добавляем ответ AI в историю
    messages.append({'role': 'assistant', 'content': reply})
    
    # Генерируем название чата из первого сообщения
    chat_title = messages[0]['content'][:50] + '...' if len(messages[0]['content']) > 50 else messages[0]['content']
    
    # Сохраняем обновлённую историю
    save_chat_history(user_email, chat_id, messages, chat_title)
    
    response_body = {
        'success': True,
        'reply': reply,
        'chat_id': chat_id
    }
    
    if thinking_text:
        response_body['thinking'] = thinking_text
    
    return {
        'statusCode': 200,
        'headers': {
            'Access-Control-Allow-Origin': '*',
            'Content-Type': 'application/json'
        },
        'body': json.dumps(response_body, ensure_ascii=False),
        'isBase64Encoded': False
    }